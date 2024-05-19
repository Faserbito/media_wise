import openpyxl

def start(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Segment_num'
    sheet['B1'] = 'Advertisement ID'
    workbook.save(path)

def work(path, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    left = [sheet['A'+str(row)].value for row in range (1, sheet.max_row+1) if (sheet['A'+str(row)].value)!=None]
    right = [sheet['B'+str(row)].value for row in range (1, sheet.max_row+1) if (sheet['B'+str(row)].value)!=None]
    sheet.cell(row = len(left) + 1, column = 1, value = data[0])
    sheet.cell(row = len(right) + 1, column = 2, value = data[1])
    workbook.save(path)

def diagram(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    #left = [sheet['A'+str(row)].value for row in range (1, sheet.max_row+1) if (sheet['A'+str(row)].value)!=None]
    right = [sheet['D'+str(row)].value for row in range (1, sheet.max_row+1) if (sheet['D'+str(row)].value)!=None]
    for i in range(3, len(right)):
        sheet.cell(row = i, column = 4, value = f"=D{i+1}-D{i}")
    #if data[0] != "":
    #    sheet.cell(row = len(left) + 1, column = 1, value = data[0])
    #if data[1] != "":
    #    sheet.cell(row = len(left) + 1, column = 2, value = data[1])
    #if data[2] != "":
    #    sheet.cell(row = len(right) + 1, column = 4, value = data[2])
    #if data[3] != "":
    #    sheet.cell(row = len(right) + 1, column = 5, value = data[3])
    workbook.save(path)